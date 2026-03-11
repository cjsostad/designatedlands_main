"""
date_filter.py — Standalone module for date-based filtering of designated lands.

Reads the `date_filter_query` column from sources_designations.csv and builds
CQL filter expressions that restrict each BCGW layer to features added, changed,
or amended within a user-specified date window.

Standalone usage:
    python date_filter.py                           # Report with default dates
    python date_filter.py --start 2025-04-01        # Custom start date
    python date_filter.py --start 2025-04-01 --end 2026-03-31
    python date_filter.py --test                    # Validate CQL against WFS

Pipeline usage (imported by designatedlands.py):
    from date_filter import build_date_filters, apply_date_filter_to_query
"""

import argparse
import csv
import os
import sys
from datetime import date, datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

requests.packages.urllib3.disable_warnings()

WFS_URL = "https://openmaps.gov.bc.ca/geo/pub/wfs"

# Sentinel values written into the CSV date_filter_query column
NO_DATE_FIELD = "no date field available"
NON_BCGW = "non-BCGW source - excluded when date filter active"

DEFAULT_START_DATE = "2025-04-01"


# ---------------------------------------------------------------------------
# Core functions (importable by the pipeline)
# ---------------------------------------------------------------------------

def build_date_filters(csv_path, start_date, end_date=None):
    """
    Read sources_designations.csv and return a list of dicts describing
    each source's date filter status.

    Parameters
    ----------
    csv_path : str
        Path to sources_designations.csv.
    start_date : str
        ISO date string (YYYY-MM-DD) for the start of the filter window.
    end_date : str or None
        ISO date string for the end of the window.  Defaults to today.

    Returns
    -------
    list[dict]
        Each dict has keys:
          - name, designation, process_order
          - bcgw_layer_name
          - date_filter_query_template  (raw template from CSV)
          - date_filter_cql             (resolved CQL, or None if excluded)
          - included                    (bool — True if layer passes filter)
          - exclude_reason              (str or None)
    """
    if end_date is None:
        end_date = date.today().isoformat()

    results = []
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("exclude", "").strip() == "T":
                continue

            name = row.get("name", "").strip()
            designation = row.get("designation", "").strip()
            process_order = row.get("process_order", "").strip()
            bcgw = row.get("bcgw_layer_name", "").strip()
            template = row.get("date_filter_query", "").strip()

            entry = {
                "name": name,
                "designation": designation,
                "process_order": process_order,
                "bcgw_layer_name": bcgw,
                "date_filter_query_template": template,
                "date_filter_cql": None,
                "included": False,
                "exclude_reason": None,
            }

            if not template or template == NO_DATE_FIELD:
                entry["exclude_reason"] = "no date field available"
            elif template == NON_BCGW:
                entry["exclude_reason"] = "non-BCGW source"
            else:
                # Resolve placeholders
                try:
                    resolved = template.format(
                        start_date=start_date,
                        end_date=end_date,
                    )
                    entry["date_filter_cql"] = resolved
                    entry["included"] = True
                except KeyError as exc:
                    entry["exclude_reason"] = f"bad placeholder: {exc}"

            results.append(entry)

    return results


def apply_date_filter_to_query(existing_query, date_filter_cql):
    """
    Combine an existing CQL/SQL query with a date filter expression.

    Parameters
    ----------
    existing_query : str or None
        The current query/CQL filter (may be empty or None).
    date_filter_cql : str
        The date filter CQL expression to append.

    Returns
    -------
    str
        Combined query with AND logic.
    """
    existing = (existing_query or "").strip()
    date_cql = (date_filter_cql or "").strip()

    if not date_cql:
        return existing
    if not existing:
        return date_cql

    return f"({existing}) AND ({date_cql})"


def get_excluded_designations(csv_path):
    """
    Return a list of designation names that cannot be date-filtered
    (no date field or non-BCGW source).
    """
    excluded = []
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("exclude", "").strip() == "T":
                continue
            template = row.get("date_filter_query", "").strip()
            if not template or template in (NO_DATE_FIELD, NON_BCGW):
                excluded.append(row.get("designation", "").strip())
    return excluded


# ---------------------------------------------------------------------------
# WFS validation helpers
# ---------------------------------------------------------------------------

def _count_features(layer, cql):
    """Query WFS with count=1 and return (http_status, feature_count_or_error)."""
    params = {
        "SERVICE": "WFS",
        "VERSION": "2.0.0",
        "REQUEST": "GetFeature",
        "typeName": layer,
        "outputFormat": "application/json",
        "CQL_FILTER": cql,
        "count": 1,
    }
    try:
        r = requests.get(WFS_URL, params=params, verify=False, timeout=60)
        if r.status_code == 200:
            data = r.json()
            return 200, data.get("numberMatched", len(data.get("features", [])))
        return r.status_code, r.text[:200]
    except requests.RequestException as exc:
        return 0, str(exc)


def _query_features(layer, cql, max_features=10000):
    """
    Query WFS and return the full list of features (as dicts of properties).
    Returns (http_status, list_of_property_dicts | error_string).
    """
    params = {
        "SERVICE": "WFS",
        "VERSION": "2.0.0",
        "REQUEST": "GetFeature",
        "typeName": layer,
        "outputFormat": "application/json",
        "CQL_FILTER": cql,
        "count": max_features,
    }
    try:
        r = requests.get(WFS_URL, params=params, verify=False, timeout=300)
        if r.status_code == 200:
            data = r.json()
            return 200, [f.get("properties", {}) for f in data.get("features", [])]
        return r.status_code, r.text[:300]
    except requests.RequestException as exc:
        return 0, str(exc)


# Map date field names to human-readable status labels
_DATE_FIELD_STATUS = {
    "ESTABLISHMENT_DATE": "New area established",
    "SECUREMENT_DATE": "Newly secured",
    "EFFECTIVE_DATE": "Newly effective",
    "DATE_OF_NOTICE": "New notice issued",
    "NOTICE_DATE": "New notice issued",
    "LEGALIZATION_DATE": "Newly legalized",
    "LAST_AMENDMENT_DATE": "Amended",
    "ORIGINAL_DECISION_DATE": "New decision",
    "LEGALIZATION_FRPA_DATE": "Legalized (FRPA)",
    "LEGALIZATION_OGAA_DATE": "Legalized (OGAA)",
    "LEGALIZATION_LAST_AMEND_DATE": "Legalization amended",
    "UPDATE_DATE": "Updated",
    "PROJECT_ESTABLISHED_DATE": "Newly established",
    "CW_DATE_CREATED": "Newly created",
}


def _extract_date_fields_from_template(template):
    """
    Parse the date_filter_query template to extract the date field names used.
    Returns a list of field names in the order they appear.
    """
    import re
    # Match field names that appear before >= '{start_date}'
    return re.findall(r"([A-Z_]+)\s*>=\s*'\{start_date\}'", template)


def _determine_status_and_date(props, date_fields, start_date):
    """
    Given feature properties and the list of date fields, determine which
    date field triggered the match and return (status_label, date_value).
    """
    for field in date_fields:
        val = props.get(field)
        if val and str(val) >= start_date:
            status = _DATE_FIELD_STATUS.get(field, "Changed")
            return status, val
    # Fallback: return the first non-null date
    for field in date_fields:
        val = props.get(field)
        if val:
            return _DATE_FIELD_STATUS.get(field, "Changed"), val
    return "Changed", ""


def _format_date(date_str):
    """Convert ISO date (YYYY-MM-DD or datetime) to MM/DD/YYYY display format."""
    if not date_str:
        return ""
    # Handle datetime strings like '2025-06-15T00:00:00Z' or '2025-06-15Z'
    clean = str(date_str).split("T")[0].split(" ")[0].rstrip("Z")
    try:
        dt = datetime.strptime(clean, "%Y-%m-%d")
        return dt.strftime("%m/%d/%Y")
    except ValueError:
        return clean


def _build_fallback_name(props, layer_name, counter):
    """
    Build a descriptive feature name from layer-specific identifier fields
    when no source_name_col or common name column is available.
    """
    # UWR layers — use UWR number + species
    uwr_num = props.get("UWR_NUMBER")
    if uwr_num:
        species = props.get("SPECIES_1", "")
        notes = props.get("FEATURE_NOTES", "")
        parts = [f"UWR {uwr_num}"]
        if species:
            parts.append(f"({species})")
        if notes:
            parts.append(f"- {notes}")
        return " ".join(parts)

    # WHA layers — use TAG + species
    tag = props.get("TAG")
    if tag:
        species = props.get("COMMON_SPECIES_NAME", "")
        parts = [f"WHA {tag}"]
        if species:
            parts.append(f"({species})")
        return " ".join(parts)

    # OGMA layers — use PROVID
    provid = props.get("LEGAL_OGMA_PROVID")
    if provid:
        return f"OGMA {provid}"

    # Generic fallback: layer name + sequential number
    return f"{layer_name} #{counter}"


def _query_feature_count(layer, existing_query, date_filter_cql):
    """
    Query WFS for a layer with the combined filter (existing + date)
    and return the matched feature count.
    """
    combined = apply_date_filter_to_query(existing_query, date_filter_cql)
    return _count_features(layer, combined)


# ---------------------------------------------------------------------------
# Excel report writer
# ---------------------------------------------------------------------------

def _unique_xlsx_path(base_path):
    """
    Return *base_path* if it does not exist yet.
    Otherwise append today's date (and a counter if needed) before .xlsx.
    """
    if not os.path.exists(base_path):
        return base_path
    root, ext = os.path.splitext(base_path)
    stamped = f"{root}_{date.today().isoformat()}{ext}"
    if not os.path.exists(stamped):
        return stamped
    counter = 2
    while True:
        candidate = f"{root}_{date.today().isoformat()}_{counter}{ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def write_report_xlsx(change_rows, excluded_entries, summary, out_path,
                      avoid_overwrite=False, federal_excluded=None,
                      pipeline_options=None):
    """
    Write the pipeline report to an Excel workbook.

    Parameters
    ----------
    change_rows : list[dict]
        Each dict has keys: layer_name, feature_name, status, date_of_change.
    excluded_entries : list[dict]
        Each dict has keys: name, exclude_reason.
    summary : dict
        Keys: start_date, end_date, total_datasets, date_filtered,
        excluded, layers_with_changes, layers_no_changes,
        total_features_changed, errors.
    out_path : str
        Desired output xlsx path.
    avoid_overwrite : bool
        If True and *out_path* already exists, append the date to the name.
    federal_excluded : list[dict] or None
        Each dict has keys: name, designation. Federal layers removed.
    pipeline_options : dict or None
        Keys describing the pipeline settings chosen by the user.

    Returns
    -------
    str
        The actual path written (may differ from *out_path* if renamed).
    """
    if avoid_overwrite:
        out_path = _unique_xlsx_path(out_path)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    wb = Workbook()

    # ---- Styles ----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="B4C6E7"),
    )
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="4472C4")

    # ==================== Changes sheet ====================
    ws = wb.active
    ws.title = "Changes"

    # Title rows
    ws.append(["Designated Lands — Pipeline Report"])
    ws["A1"].font = title_font
    ws.append([f"Date window: {summary['start_date']}  to  {summary['end_date']}"])
    ws["A2"].font = subtitle_font
    ws.append([])  # blank row

    # Column headers
    headers = ["Layer", "Feature Name", "Status", "Date of Change"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for row in change_rows:
        ws.append([
            row["layer_name"],
            row["feature_name"],
            row["status"],
            row["date_of_change"],
        ])

    # Auto-fit column widths (approximate)
    col_widths = [45, 60, 28, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[
            ws.cell(row=1, column=i).column_letter
        ].width = w

    # Light bottom border on data rows
    for r in range(5, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = thin_border

    # ==================== Excluded sheet ====================
    ws2 = wb.create_sheet("Excluded Layers")
    ws2.append(["Layer", "Reason"])
    for col_idx in (1, 2):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    for entry in excluded_entries:
        ws2.append([entry["name"], entry["exclude_reason"] or "unknown"])
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 35

    # ==================== Summary sheet ====================
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Designated Lands — Summary"])
    ws3["A1"].font = title_font
    ws3.append([])
    summary_rows = [
        ("Date window", f"{summary['start_date']}  to  {summary['end_date']}"),
        ("Total datasets", summary["total_datasets"]),
        ("Date-filtered", summary["date_filtered"]),
        ("Excluded (no date filter)", summary["excluded"]),
    ]
    if federal_excluded:
        summary_rows.append(
            ("Excluded (federal)", len(federal_excluded))
        )
    summary_rows += [
        ("Layers with changes", summary["layers_with_changes"]),
        ("Layers with no changes", summary["layers_no_changes"]),
        ("Total features changed", summary["total_features_changed"]),
    ]
    if summary.get("errors"):
        summary_rows.append(("Errors", len(summary["errors"])))
        for e in summary["errors"]:
            summary_rows.append(("", f"  - {e}"))
    for label, value in summary_rows:
        ws3.append([label, value])
        if label:
            ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True)
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 50

    # ==================== Pipeline Options sheet ====================
    ws4 = wb.create_sheet("Pipeline Options")
    ws4.append(["Designated Lands — Pipeline Options"])
    ws4["A1"].font = title_font
    ws4.append([])

    opts = pipeline_options or {}
    option_rows = [
        ("Date filtering (recent only)",
         "Enabled" if opts.get("recent_only") else "Disabled"),
        ("Start date", opts.get("start_date", summary.get("start_date", ""))),
        ("End date", opts.get("end_date", summary.get("end_date", ""))),
        ("Exclude federal layers",
         "Yes" if opts.get("exclude_federal") else "No"),
    ]
    for label, value in option_rows:
        ws4.append([label, value])
        ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True)
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 50

    # Federal excluded layers list
    if federal_excluded:
        ws4.append([])
        ws4.append(["Federal Layers Excluded from Output"])
        ws4.cell(row=ws4.max_row, column=1).font = subtitle_font
        ws4.append(["Layer Name", "Designation"])
        row_num = ws4.max_row
        for col_idx in (1, 2):
            cell = ws4.cell(row=row_num, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for fed in federal_excluded:
            ws4.append([fed["name"], fed["designation"]])

    # ==================== Designation Categories sheet ====================
    ws5 = wb.create_sheet("Designation Categories")

    cat_header_font = Font(bold=True, color="FFFFFF", size=11)
    cat_header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                                  fill_type="solid")
    cat_thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cat_fills = {
        "PPA": PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                           fill_type="solid"),
        "Protected Other": PatternFill(start_color="BDD7EE",
                                       end_color="BDD7EE",
                                       fill_type="solid"),
        "Exclude 1_2 Activities": PatternFill(start_color="FCE4D6",
                                              end_color="FCE4D6",
                                              fill_type="solid"),
        "Managed": PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                               fill_type="solid"),
    }

    def _categorize(fr, og, mr):
        fr, og, mr = fr.strip().upper(), og.strip().upper(), mr.strip().upper()
        vals = [fr, og, mr]
        if all(v == "PROTECTED" for v in vals):
            return "PPA"
        if all(v in ("FULL", "HIGH", "PROTECTED") for v in vals):
            return "Protected Other"
        if (any(v in ("FULL", "HIGH", "PROTECTED") for v in vals)
                and any(v in ("NONE", "LOW", "MEDIUM") for v in vals)):
            return "Exclude 1_2 Activities"
        return "Managed"

    cat_headers = [
        "Process Order", "Designation", "Name",
        "Forest Restriction", "OG Restriction", "Mine Restriction",
        "Category",
    ]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws5.cell(row=1, column=col_idx, value=h)
        cell.font = cat_header_font
        cell.fill = cat_header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = cat_thin_border

    csv_path = _find_csv()
    src_rows = []
    with open(csv_path, encoding="utf-8-sig") as csvf:
        for row in csv.DictReader(csvf):
            if row.get("exclude", "").strip().upper() == "T":
                continue
            src_rows.append(row)
    src_rows.sort(key=lambda x: int(x["process_order"]))

    for i, src in enumerate(src_rows, start=2):
        fr = src.get("forest_restriction", "").strip()
        og = src.get("og_restriction", "").strip()
        mr = src.get("mine_restriction", "").strip()
        cat = _categorize(fr, og, mr)
        values = [
            int(src["process_order"]),
            src.get("designation", "").strip(),
            src.get("name", "").strip(),
            fr, og, mr, cat,
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws5.cell(row=i, column=col_idx, value=v)
            cell.border = cat_thin_border
            if col_idx == 7:
                cell.fill = cat_fills.get(cat, PatternFill())
                cell.font = Font(bold=True)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    ws5.column_dimensions["A"].width = 14
    ws5.column_dimensions["B"].width = 35
    ws5.column_dimensions["C"].width = 50
    ws5.column_dimensions["D"].width = 20
    ws5.column_dimensions["E"].width = 18
    ws5.column_dimensions["F"].width = 18
    ws5.column_dimensions["G"].width = 24
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = f"A1:G{len(src_rows) + 1}"

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Standalone report
# ---------------------------------------------------------------------------

def _find_csv():
    """Locate sources_designations.csv relative to this script."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(script_dir, "sources_designations.csv")
    if os.path.exists(path):
        return path
    raise FileNotFoundError(
        f"Could not find sources_designations.csv in {script_dir}"
    )


def run_report(start_date, end_date, test_wfs=False, xlsx_path=None,
               avoid_overwrite=False, exclude_federal=False,
               federal_excluded=None, pipeline_options=None):
    """
    Print a report showing which layers have changes in the date window
    and optionally write it to an Excel spreadsheet.

    Parameters
    ----------
    start_date : str
        ISO date (YYYY-MM-DD) for the start of the filter window.
    end_date : str
        ISO date (YYYY-MM-DD) for the end of the window.
    test_wfs : bool
        If True, only validate CQL without fetching full feature details.
    xlsx_path : str or None
        If provided, write the report to this .xlsx path.
    avoid_overwrite : bool
        If True and *xlsx_path* already exists, append the date to the name.
    exclude_federal : bool
        If True, skip federal layers and record them as excluded.
    federal_excluded : list[dict] or None
        Pre-built list of federal exclusions (from the pipeline).
        If None and *exclude_federal* is True, built from the CSV.
    pipeline_options : dict or None
        Pipeline settings to record in the xlsx.

    Returns
    -------
    str or None
        Path to the xlsx file written, or None if xlsx_path was not given.
    """
    csv_path = _find_csv()
    filters = build_date_filters(csv_path, start_date, end_date)

    # Read the raw CSV to get existing queries and name columns
    source_info = {}
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            des = row.get("designation", "").strip()
            source_info[des] = {
                "query": row.get("query", "").strip(),
                "source_name_col": row.get("source_name_col", "").strip(),
                "date_filter_query": row.get("date_filter_query", "").strip(),
            }

    included = [f for f in filters if f["included"]]
    excluded = [f for f in filters if not f["included"]]

    # Handle federal exclusion in standalone mode
    if federal_excluded is None:
        federal_excluded = []
    if exclude_federal and not federal_excluded:
        # Build from CSV when running standalone
        with open(csv_path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("exclude", "").strip() == "T":
                    continue
                if row.get("jurisdiction", "").strip().lower() == "federal":
                    federal_excluded.append({
                        "name": row.get("name", "").strip(),
                        "designation": row.get("designation", "").strip(),
                    })

    # Remove federal layers from included list when excluding
    if exclude_federal and federal_excluded:
        fed_names = {f["name"] for f in federal_excluded}
        included = [f for f in included if f["name"] not in fed_names]
        excluded = [f for f in excluded if f["name"] not in fed_names]

    print("=" * 100)
    print(f"  DESIGNATED LANDS — PIPELINE REPORT")
    print(f"  Date window: {start_date}  to  {end_date}")
    if exclude_federal:
        print(f"  Federal layers: EXCLUDED ({len(federal_excluded)} layers)")
    print("=" * 100)

    total_changes = 0
    layers_with_changes = 0
    layers_no_changes = 0
    errors = []
    change_rows = []  # collected for xlsx

    for entry in included:
        layer = entry["bcgw_layer_name"]
        if not layer:
            continue

        desig = entry["designation"]
        info = source_info.get(desig, {})
        existing_q = info.get("query", "")
        name_col = info.get("source_name_col", "")
        template = info.get("date_filter_query", "")

        # Build combined CQL
        combined = apply_date_filter_to_query(existing_q, entry["date_filter_cql"])

        # Extract which date fields this layer uses
        date_fields = _extract_date_fields_from_template(template)

        if test_wfs:
            # Quick validation only
            status, count = _count_features(layer, combined)
            if status == 200:
                print(f"\n  {entry['name']}: {count} features matched (CQL OK)")
            else:
                print(f"\n  {entry['name']}: FAILED ({status}) — {count}")
                errors.append(entry["name"])
            continue

        # Full query — get actual features
        print(f"\n  {'—' * 96}")
        print(f"  {entry['name']}")
        print(f"  {'—' * 96}")

        status, result = _query_features(layer, combined)

        if status != 200:
            print(f"    ERROR querying WFS: {result}")
            errors.append(entry["name"])
            continue

        if not result:
            print(f"    No changes found in this date window.")
            layers_no_changes += 1
            continue

        layers_with_changes += 1

        # Print header
        print(f"  {'Feature Name':<55} {'Status':<25} {'Date of Change':<14}")
        print(f"  {'·' * 55} {'·' * 25} {'·' * 14}")

        feat_counter = 0
        for props in result:
            feat_counter += 1
            # Get feature name
            feature_name = ""
            if name_col:
                feature_name = str(props.get(name_col, "") or "")
            if not feature_name:
                # Fallback: try common name fields
                for fallback in ("NAME", "PROTECTED_LANDS_NAME", "SITE_NAME",
                                 "DESIGNATED_AREA_NAME", "PROJECT_NAME"):
                    feature_name = str(props.get(fallback, "") or "")
                    if feature_name:
                        break
            if not feature_name:
                # Build a descriptive name from layer-specific identifier fields
                feature_name = _build_fallback_name(props, entry["name"], feat_counter)

            # Determine status and date
            change_status, raw_date = _determine_status_and_date(
                props, date_fields, start_date,
            )
            display_date = _format_date(raw_date)

            # Truncate long names for display
            display_name = feature_name[:54]
            print(f"  {display_name:<55} {change_status:<25} {display_date:<14}")
            total_changes += 1

            change_rows.append({
                "layer_name": entry["name"],
                "feature_name": feature_name,
                "status": change_status,
                "date_of_change": display_date,
            })

    # --- Excluded layers ---
    print(f"\n  {'=' * 96}")
    print(f"  EXCLUDED LAYERS ({len(excluded)} datasets — no date filter available):\n")
    for entry in excluded:
        reason = entry["exclude_reason"] or "unknown"
        print(f"    {entry['name']:<55} {reason}")

    # --- Federal exclusions ---
    if federal_excluded:
        print(f"\n  {'=' * 96}")
        print(f"  FEDERAL LAYERS EXCLUDED ({len(federal_excluded)} datasets):\n")
        for fed in federal_excluded:
            print(f"    {fed['name']:<55} {fed['designation']}")

    # --- Summary ---
    print(f"\n{'=' * 100}")
    print(f"  SUMMARY")
    print(f"    Date window:             {start_date}  to  {end_date}")
    print(f"    Exclude federal:         {'Yes' if exclude_federal else 'No'}")
    print(f"    Total datasets:          {len(filters)}")
    print(f"    Date-filtered:           {len(included)}")
    print(f"    Excluded (no date):      {len(excluded)}")
    if federal_excluded:
        print(f"    Excluded (federal):      {len(federal_excluded)}")
    if not test_wfs:
        print(f"    Layers with changes:     {layers_with_changes}")
        print(f"    Layers with no changes:  {layers_no_changes}")
        print(f"    Total features changed:  {total_changes}")
    if errors:
        print(f"    Errors:                  {len(errors)}")
        for e in errors:
            print(f"      - {e}")
    print(f"{'=' * 100}\n")

    # --- Write xlsx ---
    if xlsx_path and not test_wfs:
        if pipeline_options is None:
            pipeline_options = {
                "recent_only": True,
                "exclude_federal": exclude_federal,
                "start_date": start_date,
                "end_date": end_date,
            }
        summary = {
            "start_date": start_date,
            "end_date": end_date,
            "total_datasets": len(filters),
            "date_filtered": len(included),
            "excluded": len(excluded),
            "layers_with_changes": layers_with_changes,
            "layers_no_changes": layers_no_changes,
            "total_features_changed": total_changes,
            "errors": errors,
        }
        excluded_info = [
            {"name": e["name"], "exclude_reason": e["exclude_reason"]}
            for e in excluded
        ]
        written = write_report_xlsx(
            change_rows, excluded_info, summary, xlsx_path,
            avoid_overwrite=avoid_overwrite,
            federal_excluded=federal_excluded,
            pipeline_options=pipeline_options,
        )
        print(f"  Report saved to: {written}\n")
        return written

    return None


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Date filter module for designated lands. "
            "Run standalone to produce a report showing which layers "
            "had changes within the specified date window."
        ),
    )
    parser.add_argument(
        "--start",
        metavar="YYYY-MM-DD",
        default=DEFAULT_START_DATE,
        help=f"Start date for the filter window (default: {DEFAULT_START_DATE})",
    )
    parser.add_argument(
        "--end",
        metavar="YYYY-MM-DD",
        default=None,
        help="End date for the filter window (default: today)",
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Validate each CQL filter against the WFS endpoint and report feature counts",
    )
    parser.add_argument(
        "--output", "-o",
        metavar="PATH",
        default=None,
        help="Output path for the .xlsx report (default: outputs/designated_lands_pipeline_report.xlsx)",
    )
    parser.add_argument(
        "--exclude-federal",
        action="store_true",
        help="Exclude federally protected areas from the report",
    )
    args = parser.parse_args()

    # Validate date formats
    for label, val in [("start", args.start), ("end", args.end)]:
        if val is not None:
            try:
                datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                print(f"ERROR: {label} date '{val}' is not valid YYYY-MM-DD format")
                sys.exit(1)

    end_date = args.end or date.today().isoformat()

    # Default xlsx output path (next to this script)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_xlsx = os.path.join(
        script_dir, "outputs", "designated_lands_pipeline_report.xlsx",
    )
    xlsx_out = getattr(args, "output", None) or default_xlsx

    run_report(args.start, end_date, test_wfs=args.test, xlsx_path=xlsx_out,
               exclude_federal=args.exclude_federal)


if __name__ == "__main__":
    main()
